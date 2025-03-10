
# -*- coding: utf-8 -*-
"""
Created on Fri Sep  6 12:33:12 2024

@author: inarisetty
"""
import pandas as pd
from openpyxl import load_workbook


def specs_transform(excel_file):
    # Load the workbook
    workbook = load_workbook(excel_file)

    # Check for the presence of 'Study' sheet
    if 'Study' in workbook.sheetnames:
        study_sheet = workbook['Study']
        study_data = study_sheet.values
        study_columns = next(study_data)
        study_df = pd.DataFrame(study_data, columns=study_columns)
        # Select only the required columns
        study_info = study_df.loc[study_df['Attribute'].isin([
            'Project', 'Study Name', 'Study Description', 
            'Protocol Number', 'Metadata Name', 'Metadata Description', 
            'MedDRA Dictionary Version', 'WHODrug Version'
        ])][['Attribute', 'Value']]
    else:
        study_info = None  # Set to None if the sheet is missing
        print("Warning: The 'Study' sheet is missing in the provided Excel file.")

    # Check for the presence of 'Standards' sheet
    if 'Standards' in workbook.sheetnames:
        standards_sheet = workbook['Standards']
        standards_data = standards_sheet.values
        standards_columns = next(standards_data)
        standards_df = pd.DataFrame(standards_data, columns=standards_columns)
    else:
        standards_df = None  # Set to None if the sheet is missing
        print("Warning: The 'Standards' sheet is missing in the provided Excel file.")

    # Handle the TOC and combine other sheets
    toc_sheet_name = 'TOC'
    if toc_sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{toc_sheet_name}' not found in the Excel file.")
    
    toc_sheet = workbook[toc_sheet_name]
    toc_data = toc_sheet.values
    toc_columns = next(toc_data)
    toc_df = pd.DataFrame(toc_data, columns=toc_columns)
    
    # Filter the TOC DataFrame for Active = "Y"
    active_datasets = toc_df[toc_df['Active'] == 'Y']['Dataset'].unique()

    # Define the specific datasets that should be stacked last
    special_datasets = ['TA', 'TD', 'TV', 'TE', 'TM', 'TS', 'TI']

    # Function to convert a sheet to a DataFrame and add a 'Sheet' column
    def sheet_to_dataframe(sheet, sheet_name):
        data = sheet.values
        columns = next(data)
        df = pd.DataFrame(data, columns=columns)
        df.insert(0, 'Sheet', sheet_name)  # Add the 'Sheet' column
        return df

    # Read all active sheets and store them in lists
    df_list = []
    special_df_list = []

    for sheet_name in active_datasets:
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            df = sheet_to_dataframe(sheet, sheet_name)
            if sheet_name in special_datasets:
                special_df_list.append(df)
            else:
                df_list.append(df)
        else:
            print(f"Warning: Sheet '{sheet_name}' listed in TOC not found in the Excel file.")

    # Keep only the specified columns in each DataFrame
    columns_to_keep = ['Sheet', 'Dataset', 'Label', 'Class', 'Display Order', 'Input Datasets', 'Variable', 'Input Variables', 'Mapping Action', 'Implemented SAS Code', 'Mapping Rule']

    def filter_columns(df):
        available_columns = [col for col in columns_to_keep if col in df.columns]
        return df[available_columns]

    df_list = [filter_columns(df) for df in df_list]
    special_df_list = [filter_columns(df) for df in special_df_list]

    # Reset index before concatenation
    df_list = [df.reset_index(drop=True) for df in df_list]
    special_df_list = [df.reset_index(drop=True) for df in special_df_list]

    # Concatenate all DataFrames, stacking special datasets last
    combined_df = pd.concat(df_list + special_df_list, ignore_index=True)

    # Drop the 'Display Format' column from combined_df if it exists
    if 'Display Format' in combined_df.columns:
        combined_df = combined_df.drop(columns=['Display Format'])

    # Select relevant columns from TOC for merging
    toc_columns_to_merge = toc_df[['Dataset', 'Label', 'Class', 'Display Order', 'Input Datasets']]
    toc_columns_to_merge = toc_columns_to_merge.rename(columns={'Dataset': 'Sheet'})

    # Merge combined_df with the TOC columns
    combined_df = combined_df.merge(toc_columns_to_merge, on='Sheet', how='left')
    combined_df.rename(columns={
        'Label_x': 'Label',
        'Display Order_x': 'Display Order',
        'Label_y': 'Dataset Label',
        'Display Order_y': 'Dataset Order'
    }, inplace=True)

    # Filter rows where 'Variable' is not null and not empty after stripping
    combined_df = combined_df[combined_df['Variable'].notnull() & (combined_df['Variable'].str.strip() != '')]

    return study_info, standards_df, combined_df

