# -*- coding: utf-8 -*-
"""
Created on Wed Jul 31 10:32:36 2024

@author: inarisetty
"""

import pandas as pd
import plotly.express as px
from openpyxl import load_workbook
from plotly.io import write_image
import plotly.graph_objects as go
from plotly.subplots import make_subplots

# Load the workbook and TOC sheet
excel_file = 'G:/projects/p592/s5926173/csdtm_dev/draft1/docs/sdtm/s5926173_csdtm_dev_sdtm_mapping.xlsx' 
workbook = load_workbook(excel_file)
toc_sheet_name = 'TOC'

if toc_sheet_name not in workbook.sheetnames:
    raise ValueError(f"Sheet '{toc_sheet_name}' not found in the Excel file.")

# Convert the TOC sheet to a DataFrame
toc_sheet = workbook[toc_sheet_name]
toc_data = toc_sheet.values
toc_columns = next(toc_data)[0:]
toc_df = pd.DataFrame(toc_data, columns=toc_columns)

# Check if the 'Active' column exists
if 'Active' not in toc_df.columns:
    raise ValueError("Column 'Active' not found in the TOC sheet.")
unique_values_counts = toc_df['Active'].value_counts(dropna=False).reset_index()
unique_values_counts.columns = ['Active', 'Count']

#unique_values_counts['Active'] = unique_values_counts['Active'].fillna('')

# Calculate percentage
total_count = unique_values_counts['Count'].sum()
unique_values_counts['Percentage'] = (unique_values_counts['Count'] / total_count) * 100

# Create labels for the bars
unique_values_counts['Label'] = unique_values_counts.apply(
    lambda x: f"Count: {x['Count']}<br>Percentage: {x['Percentage']:.2f}%", axis=1)

# Create a bar plot using Plotly with different colors for each bar
fig_active = px.bar(unique_values_counts, x='Active', y='Count', title='Unique Values in Active Column',
             text='Label', labels={'Active': 'Active', 'Count': 'Count'}, height=800)

# =============================================================================
# # Calculate the number of unique values in the 'Active' column
# unique_values_counts = toc_df['Active'].value_counts(dropna=False).reset_index()
# unique_values_counts.columns = ['Active', 'Count']
# unique_values_counts['Active'] = unique_values_counts['Active'].fillna('')
# 
# # Ensure all expected values are included
# expected_values = ['Y', 'N', '']
# for value in expected_values:
#     if value not in unique_values_counts['Active'].values:
#         unique_values_counts = pd.concat([unique_values_counts, pd.DataFrame({'Active': [value], 'Count': [0]})], ignore_index=True)
# 
# # Calculate percentage
# total_count = unique_values_counts['Count'].sum()
# unique_values_counts['Percentage'] = (unique_values_counts['Count'] / total_count) * 100
# 
# # Create labels for the bars
# unique_values_counts['Label'] = unique_values_counts.apply(
#     lambda x: f"Count: {x['Count']}<br>Percentage: {x['Percentage']:.2f}%", axis=1)
# 
# # Create a bar plot for Active column
# fig_active = px.bar(unique_values_counts, x='Active', y='Count', title='Unique Values in Active Column',
#                     text='Label', color='Active', labels={'Active': 'Active', 'Count': 'Count'}, height=600)
# =============================================================================
fig_active.update_traces(texttemplate='%{text}', textposition='outside')
fig_active.update_layout(legend_title_text='Active')

# Create pie chart for Datasets column
if 'Dataset' not in toc_df.columns:
    raise ValueError("Column 'Dataset' not found in the TOC sheet.")

# Calculate the number of unique values in the Datasets column
datasets_counts = toc_df['Dataset'].value_counts(dropna=False).reset_index()
datasets_counts.columns = ['Dataset', 'Count']
datasets_counts['Dataset'] = datasets_counts['Dataset'].fillna('')

# Create a pie chart for Datasets column
#fig_datasets = px.pie(datasets_counts, names='Dataset', values='Count', title='Unique Values in Dataset Column')

# Create bar plot for Class column
if 'Class' not in toc_df.columns:
    raise ValueError("Column 'Class' not found in the TOC sheet.")

# Calculate the number of unique values in the Class column
toc_df_active = toc_df[toc_df['Active'] == 'Y']
# =============================================================================
# class_counts = toc_df_active['Class'].value_counts(dropna=False).reset_index()
# class_counts.columns = ['Class', 'Count','Dataset']
# class_counts['Class'] = class_counts['Class'].fillna('')
# class_counts['Label'] = class_counts.apply(
#     lambda x: f"Count: {x['Count']}", axis=1)
# 
# 
# # Create a grouped bar plot for Class column
# fig_class = px.bar(class_counts, x='Class', y='Dataset', color='Dataset', barmode='group',
#                    title='Unique Values in Class Column',
#                    text='Label', labels={'Class': 'Class', 'Dataset': 'Dataset'}, height=600)
# fig_class.update_traces(texttemplate='%{text}', textposition='auto')
# fig_class.update_layout(showlegend=True, legend_title_text='Dataset')
# =============================================================================
class_datasets_counts = toc_df_active.groupby(['Class']).size().reset_index(name='Count')

# Create labels for the bars
class_datasets_counts['Label'] = class_datasets_counts.apply(
    lambda x: f"Count: {x['Count']}", axis=1)

# Create a grouped bar plot for Class column
fig_class = px.bar(class_datasets_counts, x='Class', y='Count', color='Dataset', 
                   title='Unique Values in Class Column',
                   text='Label', labels={'Class': 'Class'}, height=600)
fig_class.update_traces(texttemplate='%{text}', textposition='auto')
fig_class.update_layout(showlegend=True, legend_title_text='Datasets')
# Combine all plots into a single PDF
# Create a subplot figure with landscape layout
fig_combined = make_subplots(
    rows=2, cols=1,
    subplot_titles=('Unique Values in Active Column', 'Unique Values in Class Column'),
    specs=[[{"type": "xy"}], [{"type": "xy"}]],
    vertical_spacing=0.1
)

# Add the plots to the subplot figure
fig_combined.add_trace(fig_active['data'][0], row=1, col=1)
#fig_combined.add_trace(fig_datasets['data'][0], row=2, col=1)
fig_combined.add_trace(fig_class['data'][0], row=2, col=1)

# Update layout
fig_combined.update_layout(height=1800, width=1200, showlegend=True)

# Save the combined plot as a PDF
output_pdf = 'combined_unique_values_plots.pdf'
write_image(fig_combined, output_pdf, format='pdf')

print(f"The combined plot has been saved as {output_pdf}")