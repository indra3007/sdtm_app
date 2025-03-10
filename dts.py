# -*- coding: utf-8 -*-
"""
Created on Wed Aug  7 11:24:47 2024

@author: inarisetty
"""

import pyreadstat
import pandas as pd 

import dtale
from openpyxl import load_workbook
import os

from datetime import datetime
from dm_dates import process_datasets
project = "p627"
study = "s221"
excel_file = f"G:/users/inarisetty/private/python/combined_data_{study}.xlsx"
sas_directory = f"G:/projects/{project}/{study}/csdtm_dev/draft1/rawdata/"
sdtm_directory = f"G:/projects/{project}/{study}/csdtm_dev/draft1/sdtmdata/"
workbook = load_workbook(excel_file)
sheet_name = 'Sheet1'
sheet = workbook[sheet_name]
data = sheet.values
columns = next(data)[0:]

proc_var = "RFPENDTC"
df_xl = pd.DataFrame(data, columns=columns)

# Filter the DataFrame based on the dataset and variable
df_xl = df_xl[(df_xl['Dataset'] == 'DM') & (df_xl['Variable'] == proc_var)]

# Select specific columns
df_xl = df_xl.loc[:, ['Dataset', 'Variable', 'Input Variables', 'Mapping Action', 'Implemented SAS Code']]

# Process the datasets
result_df = process_datasets(df_xl, sas_directory, aggregation='min', remove_time=False, all_dates=True, remove_partial=False)

result_df.to_excel(f'{proc_var}_CHK.xlsx', index=False)
# Create a new column for checking
result_df[f'{proc_var}_CHK'] = result_df[proc_var] 
result_df['USUBJID'] = result_df['X_SUBJID']
result_df[['part1', 'part2', 'part3', 'part4', 'part5']] = result_df['USUBJID'].str.split('-', expand=True)
result_df['USUBJID'] = result_df.apply(lambda row: f"{row['part1']}-{row['part2']}-{row['part4']}-{row['part5']}", axis=1)

# Display the DataFrame with the new `usubjid`
# Step 2: Compare the third and fourth parts
result_df['inv_Match'] = result_df.apply(lambda row: 'Match' if row['part3'] == row['part4'] else 'Mis Match', axis=1)

#result_df.to_excel(f'{proc_var}_DEV.xlsx', index=False)
# Define the path to the SDTM file
sdtm_file = os.path.join(sdtm_directory, "dm.sas7bdat")

# Read the SAS dataset
df_dm, meta = pyreadstat.read_sas7bdat(sdtm_file, encoding="LATIN1")

# Check if df_dm is a DataFrame and select relevant columns
if isinstance(df_dm, pd.DataFrame):
    df_dm = df_dm[['USUBJID', proc_var]]
    #df_dm[proc_var] = df_dm[proc_var].apply(lambda x: x.split('T')[0] if 'T' in x else x)
else:
    print("df_dm is not a DataFrame, something went wrong.")

# Merge the DataFrames and compare the variables
final_chk = pd.merge(df_dm, result_df[['USUBJID', f'{proc_var}_CHK']], on='USUBJID', how='left')
final_chk['Compare'] = final_chk.apply(
    lambda row: 'Match' if (
        (pd.isnull(row[f'{proc_var}_CHK']) and pd.isnull(row[proc_var])) or  # Both are null
        (row[f'{proc_var}_CHK'] == row[proc_var]) or  # Both are equal
        (row[f'{proc_var}_CHK'] == '' and row[proc_var] == '')  # Both are empty strings
    ) else 'Mis Match', 
    axis=1
)
final_chk = final_chk[final_chk[proc_var].notnull() & (final_chk[proc_var].str.strip() != '') &  (final_chk[f'{proc_var}_CHK'].str.strip() != '') & final_chk[f'{proc_var}_CHK'].notnull()]

# Save the final DataFrame to an Excel file
final_chk.to_excel(f'{proc_var}.xlsx', index=False)


# =============================================================================
# 
# 
# dir_path = "G:/users/inarisetty/private/python/"
# 
# excel_file = f'{proc_var}.xlsx'
# 
# def load_excel_sheet_to_dataframe(dir_path, excel_file, sheet_name='Sheet1'):
#     excel_file = os.path.join(dir_path, excel_file)
#     # Load the workbook and select the specified sheet
#     workbook = load_workbook(excel_file)
#     sheet = workbook[sheet_name]
#     
#     # Extract data and columns
#     data = sheet.values
#     columns = next(data)[0:]  # Get the first row for column names
#     
#     # Load the rest of the data into a DataFrame
#     df_xl = pd.DataFrame(data, columns=columns)
#     
#     return df_xl
# 
# df_xl_dt = load_excel_sheet_to_dataframe(dir_path, excel_file, sheet_name='Sheet1')
# df_xl_dt['Dates'] = df_xl_dt[proc_var] 
# 
# all_d_file = "all_dates.xlsx"
# df_xl_alldt = load_excel_sheet_to_dataframe(dir_path, all_d_file, sheet_name='Sheet1')
# df_xl_alldt['USUBJID'] = df_xl_alldt['X_SUBJID'] 
# 
# 
# final_chk = pd.merge(df_xl_dt, df_xl_alldt,  on=['USUBJID', 'Dates'], how='left')
# final_chk.to_excel('final_chk.xlsx', index=False)
# =============================================================================


